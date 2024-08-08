import json
from copy import copy
from functools import reduce

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.utils.decorators import method_decorator
from weasyprint import HTML

from admin_panel.tasks import *

from django.core import serializers
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse, JsonResponse, Http404, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.urls import reverse_lazy, reverse
from django.views.generic import *
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import range_boundaries

from admin_panel.forms import *
from admin_panel.models import *
import House24.settings as settings


class StaffRequiredMixin():
    """
    Mixin which requires that the authenticated user is a staff member.
    """

    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_staff:

            messages.error(
                request,
                'Либо вы не вошли в систему'
                ' либо у вас нету соответствующих прав для пользования административной панелью.'
            )
            return redirect(settings.LOGIN_URL)
        return super(StaffRequiredMixin, self).dispatch(request, *args, **kwargs)


def balances(context):
    total_plus = sum(Paybox.objects.filter(debit_credit='plus', is_complete=True).values_list('total', flat=True))
    total_minus = sum(Paybox.objects.filter(debit_credit='minus', is_complete=True).values_list('total', flat=True))
    context['paybox_balance'] = total_plus - total_minus

    context['personal_accounts_debts'] = sum(
        PersonalAccount.objects.filter(balance__lt=0).values_list('balance', flat=True))

    personal_accounts_plus = sum(
        Paybox.objects.filter(debit_credit='plus', is_complete=True, personal_account__isnull=False).values_list(
            'total', flat=True))
    personal_accounts_minus = sum(
        Receipt.objects.filter(is_complete=True, flat__personal_account__isnull=False).values_list(
            'total_price', flat=True))
    context['personal_accounts_balance'] = personal_accounts_plus - personal_accounts_minus


class Statistic(StaffRequiredMixin, TemplateView):
    template_name = 'admin_panel/statistic.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['houses_count'] = House.objects.all().count()
        context['personal_accounts_count'] = PersonalAccount.objects.all().count()
        context['flats_count'] = Flat.objects.all().count()
        context['flat_owner_count'] = FlatOwner.objects.all().count()
        context['in_work_applications_count'] = Application.objects.filter(status='in work').count()
        context['new_applications_count'] = Application.objects.filter(status='new').count()
        balances(context)

        receipt_minus_by_month = []
        receipt_plus_by_month = []
        for i in range(1, 13):
            receipt_plus = Receipt.objects.aggregate(
                sum=Sum('total_price', filter=Q(date_published__month=i) & Q(is_complete=True) & Q(status="paid")))
            receipt_minus = Receipt.objects.aggregate(
                sum=Sum('total_price', filter=Q(date_published__month=i) & Q(is_complete=True) & Q(status="unpaid")))

            receipt_minus_by_month.append(int(receipt_minus['sum']) if receipt_minus['sum'] else 0)
            receipt_plus_by_month.append(int(receipt_plus['sum']) if receipt_plus['sum'] else 0)

        context['receipt_minus_by_month'] = receipt_minus_by_month
        context['receipt_plus_by_month'] = receipt_plus_by_month

        paybox_minus_by_month = []
        paybox_plus_by_month = []
        for i in range(1, 13):
            paybox_plus = Paybox.objects.aggregate(
                sum=Sum('total', filter=Q(date_published__month=i) & Q(is_complete=True) & Q(debit_credit="plus")))
            paybox_minus = Paybox.objects.aggregate(
                sum=Sum('total', filter=Q(date_published__month=i) & Q(is_complete=True) & Q(debit_credit="minus")))

            paybox_minus_by_month.append(int(paybox_minus['sum']) if paybox_minus['sum'] else 0)
            paybox_plus_by_month.append(int(paybox_plus['sum']) if paybox_plus['sum'] else 0)

        context['paybox_minus_by_month'] = paybox_minus_by_month
        context['paybox_plus_by_month'] = paybox_plus_by_month
        return context


class PayboxList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/paybox.html'
    context_object_name = 'paybox'
    paginate_by = 20

    def get_queryset(self):
        paybox = Paybox.objects.all()
        wb = Workbook()
        ws = wb.active  # это лист в excel
        ws.append(
            [
                '#',
                'Дата',
                'Приход/расход',
                'Статус',
                'Статья',
                'Квитанция',
                'Услуга',
                'Сумма',
                'Валюта',
                'Владелец квартиры',
                'Лицевой счет',
            ]
        )
        # Define a font style that is bold
        bold_font = Font(bold=True)
        # Apply the bold font style to the cells in the row
        for cell in ws[1]:
            cell.font = bold_font

        for obj in paybox:
            if obj.flat_owner is not None:
                flat_owner = f'{obj.flat_owner}'
            else:
                flat_owner = f''

            if obj.personal_account is not None:
                personal_account = f'{obj.personal_account}'
            else:
                personal_account = f''
            ws.append([
                f'{obj.number}',
                f'{obj.date_published}',
                f'{obj.get_debit_credit_display()}',
                f'{"Проведен" if obj.is_complete else "Не проведен"}',
                f'{obj.article.title}',
                f'',
                f'',
                f'{obj.total}',
                f'UAH',
                f'{flat_owner}',
                f'{personal_account}',
            ])

        ws.title = "Выписка"  # это название листа в excel
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 20

        wb.save('media/paybox/all_info.xlsx')

        # Get full path to workbook
        wb.close()

        return paybox

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        total_plus = sum(Paybox.objects.filter(debit_credit='plus', is_complete=True).values_list('total', flat=True))
        total_minus = sum(Paybox.objects.filter(debit_credit='minus', is_complete=True).values_list('total', flat=True))
        context['filter_form'] = PayboxFilterForm()
        context['total_plus'] = total_plus
        context['total_minus'] = total_minus
        balances(context)
        return context


class PayboxFilteredList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/paybox.html'
    context_object_name = 'paybox'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = PayboxFilterForm(initial=self.request.GET)
        balances(context)
        return context

    def get_queryset(self):
        paybox = Paybox.objects.all()
        form_filter = PayboxFilterForm(self.request.GET)
        qs = []
        if form_filter.is_valid():
            if form_filter.cleaned_data['number']:
                qs.append(Q(number__icontains=form_filter.cleaned_data['number']))
            if form_filter.cleaned_data['personal_account']:
                qs.append(Q(personal_account__number__icontains=form_filter.cleaned_data['personal_account']))
            if form_filter.cleaned_data['daterange']:
                date_start, date_end = str(form_filter.cleaned_data['daterange']).split(' - ')
                date_start = date_start.split('/')
                date_end = date_end.split('/')
                date_start.reverse()
                date_end.reverse()
                date_start = "-".join(date_start)
                date_end = "-".join(date_end)
                qs.append(Q(
                    Q(date_published__gte=date_start) &
                    Q(date_published__lte=date_end)
                ))

            if form_filter.cleaned_data['article']:
                qs.append(Q(article_id=form_filter.cleaned_data['article'].id))
            if form_filter.cleaned_data['flat_owner']:
                full_name = str(form_filter.cleaned_data['flat_owner']).split(' ')
                qs.append(Q(
                    Q(flat_owner__user__last_name__icontains=full_name[0]) &
                    Q(flat_owner__patronymic__icontains=full_name[2]) &
                    Q(flat_owner__user__first_name__icontains=full_name[1])
                ))

            if form_filter.cleaned_data['status']:
                if form_filter.cleaned_data['status'] == 'complete':
                    qs.append(Q(is_complete=True))
                if form_filter.cleaned_data['status'] == 'no complete':
                    qs.append(Q(is_complete=False))
            if form_filter.cleaned_data['debit_credit']:
                qs.append(Q(debit_credit=form_filter.cleaned_data['debit_credit']))
            q = Q()
            for item in qs:
                q = q & item
            paybox = Paybox.objects.filter(q)
        return paybox


class CreatePaybox(StaffRequiredMixin, FormView):
    def get(self, request, income, *args, **kwargs):
        form = PayboxForm()
        form.fields['date_published'].initial = timezone.now().date()
        form.fields['user'].initial = Personal.objects.get(user_id=self.request.user.id)
        form.fields['is_complete'].initial = True

        if income == 'plus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        elif income == 'minus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="minus")
        data = {
            'income': income,
            'form': form,
        }
        return render(request, 'admin_panel/get_paybox_form.html', context=data)

    def post(self, request, income, *args, **kwargs):
        form = PayboxForm(request.POST)

        if form.is_valid():
            instance = form.save()
            if income == 'plus':
                if instance.personal_account is not None and instance.is_complete is True:
                    personal_account = PersonalAccount.objects.get(pk=instance.personal_account_id)
                    personal_account.balance = personal_account.balance + instance.total
                    personal_account.save()
                instance.debit_credit = 'plus'
            elif income == 'minus':
                instance.debit_credit = 'minus'
            instance.save()
            return redirect('paybox')
        else:

            data = {
                'income': income,
                'form': form,
            }
            return render(request, 'admin_panel/get_paybox_form.html', context=data)


class CopyPaybox(CreatePaybox):
    def get(self, request, pk, *args, **kwargs):
        copy = Paybox.objects.get(pk=pk)
        form = PayboxForm(instance=copy)

        if copy.debit_credit == 'plus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        elif copy.debit_credit == 'minus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="minus")
        income = copy.debit_credit
        data = {
            'income': income,
            'form': form,
        }
        return render(request, 'admin_panel/get_paybox_form.html', context=data)


from urllib.parse import urlencode, quote


class PayboxDetail(StaffRequiredMixin, DetailView):
    model = Paybox
    template_name = 'admin_panel/read_paybox.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        paybox = Paybox.objects.get(pk=self.kwargs['pk'])
        context['paybox'] = paybox

        wb = Workbook()
        ws = wb.active  # это лист в excel
        ws.append(['Платёж', f'#{paybox.number}'])
        ws.append(['Дата', f'{paybox.date_published}'])
        if paybox.flat_owner:
            ws.append(['Владелец квартиры',
                       f'{paybox.flat_owner.user.last_name} {paybox.flat_owner.user.first_name} {paybox.flat_owner.patronymic}'])
        else:
            ws.append(['Владелец квартиры', f'не указан'])

        if paybox.personal_account:
            ws.append(['Лицевой счёт', paybox.personal_account.number])
        else:
            ws.append(['Лицевой счёт', f'не указан'])

        ws.append(['Приход/Расход', paybox.get_debit_credit_display()])
        ws.append(['Проведён', 'Проведён' if paybox.is_complete else 'Не проведён'])
        if paybox.article:
            ws.append(['Статья', paybox.article.title])
        else:
            ws.append(['Статья', f'не указана'])

        ws.append(['Квитанция', ''])
        ws.append(['Услуга', ''])
        ws.append(['Сумма', f'{paybox.total}'])
        ws.append(['Валюта', 'UAH'])
        ws.append(['Комментарий', paybox.comment])
        ws.append(['Приход/Расход', paybox.get_debit_credit_display()])
        ws.append(['Проведён', 'Проведён' if paybox.is_complete else 'Не проведён'])
        if paybox.user:
            ws.append(['Менеджер', f'{paybox.user.user.last_name} {paybox.user.user.first_name}'])
        else:
            ws.append(['Менеджер', f'не указан'])
        ws.title = "Выписка"  # это название листа в excel
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

        wb.save('media/paybox/info.xlsx')

        # Get full path to workbook
        wb.close()

        return context


class UpdatePaybox(StaffRequiredMixin, FormView):
    def get(self, request, pk, *args, **kwargs):
        paybox = Paybox.objects.get(pk=pk)
        form = PayboxForm(instance=paybox)
        if paybox.debit_credit == 'plus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        elif paybox.debit_credit == 'minus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="minus")
        data = {
            'form': form,
        }
        return render(request, 'admin_panel/update_paybox.html', context=data)

    def post(self, request, pk, *args, **kwargs):
        paybox = Paybox.objects.get(pk=pk)
        form = PayboxForm(request.POST, instance=paybox)
        if form.is_valid():
            instance = form.save()
            if instance.debit_credit == 'plus':

                if instance.personal_account is not None and instance.is_complete is True:
                    personal_account = PersonalAccount.objects.get(pk=instance.personal_account_id)
                    plus_total = sum(
                        Paybox.objects.filter(personal_account=personal_account,
                                              is_complete=True).values_list('total', flat=True))
                    minus_total = sum(
                        Receipt.objects.filter(flat__personal_account=personal_account,
                                               is_complete=True).values_list('total_price', flat=True))

                    personal_account.balance = plus_total - minus_total
                    personal_account.save()
            return redirect('paybox')
        else:
            data = {
                'form': form,
            }
            return render(request, 'admin_panel/update_paybox.html', context=data)


class DeletePaybox(StaffRequiredMixin, FormView):
    def post(self, request, pk, *args, **kwargs):
        paybox = Paybox.objects.get(pk=pk)
        if paybox.debit_credit == 'plus':
            if paybox.personal_account is not None and paybox.is_complete is True:
                personal_account = PersonalAccount.objects.get(pk=paybox.personal_account_id)
                paybox.delete()
                plus_total = sum(
                    Paybox.objects.filter(personal_account=personal_account,
                                          is_complete=True).values_list('total', flat=True))
                minus_total = sum(
                    Receipt.objects.filter(flat__personal_account=personal_account,
                                           is_complete=True).values_list('total_price', flat=True))

                personal_account.balance = plus_total - minus_total
                personal_account.save()
            else:
                paybox.delete()

        else:
            paybox.delete()
        return redirect('paybox')


class ReceiptList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/receipts.html'
    context_object_name = 'receipts'
    queryset = Receipt.objects.all()
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ReceiptFilterForm()
        balances(context)
        return context


class CreateReceipt(StaffRequiredMixin, CreateView):
    model = Receipt
    template_name = 'admin_panel/get_receipt_form.html'
    form_class = ReceiptForm
    success_url = reverse_lazy('receipts')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['indications'] = Indication.objects.order_by('date_published').all()
        context['service_formset'] = ReceiptServiceFormset(queryset=ReceiptService.objects.none(), prefix='service')
        context['measures'] = Measure.objects.all()
        context['services'] = Service.objects.all()
        balances(context)
        return context

    def post(self, request, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST)
        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save(commit=False)
            for instance in instances:
                instance.receipt_id = obj.id
                instance.save()
            if hasattr(obj.flat, 'personal_account'):
                if obj.flat.personal_account is not None \
                        and obj.flat.personal_account != '' \
                        and obj.is_complete is True:
                    personal_account = PersonalAccount.objects.get(pk=obj.flat.personal_account.id)
                    personal_account.balance = personal_account.balance - obj.total_price
                    personal_account.save()
            return redirect('receipts')
        else:
            data = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, 'admin_panel/get_receipt_form.html', context=data)


class GetIndicationsSortedList(StaffRequiredMixin, View):
    def get(self, request, flat_id):
        indications = Indication.objects.order_by('date_published').filter(flat_id=flat_id)
        data = {
            "indications": indications,
        }
        return render(request, 'admin_panel/ajax_indication_table.html', context=data)


class ReceiptsFilteredList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/receipts.html'
    context_object_name = 'receipts'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ReceiptFilterForm(initial=self.request.GET)
        balances(context)
        return context

    def get_queryset(self):
        receipts = Receipt.objects.all()
        form_filter = ReceiptFilterForm(self.request.GET)
        qs = []
        if form_filter.is_valid():
            if form_filter.cleaned_data['number']:
                qs.append(Q(number__icontains=form_filter.cleaned_data['number']))
            if form_filter.cleaned_data['status']:
                qs.append(Q(status=form_filter.cleaned_data['status']))
            if form_filter.cleaned_data['daterange']:
                date_start, date_end = str(form_filter.cleaned_data['daterange']).split(' - ')
                date_start = date_start.split('/')
                date_end = date_end.split('/')
                date_start.reverse()
                date_end.reverse()
                date_start = "-".join(date_start)
                date_end = "-".join(date_end)
                qs.append(Q(
                    Q(date_published__gte=date_start) &
                    Q(date_published__lte=date_end)
                ))
            if form_filter.cleaned_data['month']:
                date_list = form_filter.cleaned_data['month'].split('-')
                date_list.reverse()
                result = "-".join(date_list)
                qs.append(Q(date_published__month=date_list[1]))
            if form_filter.cleaned_data['flat']:
                qs.append(Q(flat__number__icontains=form_filter.cleaned_data['flat']))
            if form_filter.cleaned_data['flat_owner']:
                full_name = str(form_filter.cleaned_data['flat_owner']).split(' ')
                qs.append(Q(
                    Q(flat__flat_owner__user__last_name__icontains=full_name[0]) &
                    Q(flat__flat_owner__patronymic__icontains=full_name[2]) &
                    Q(flat__flat_owner__user__first_name__icontains=full_name[1])
                ))

            if form_filter.cleaned_data['complete']:
                if form_filter.cleaned_data['complete'] == 'complete':
                    qs.append(Q(is_complete=True))
                if form_filter.cleaned_data['complete'] == 'no complete':
                    qs.append(Q(is_complete=False))
            q = Q()
            for item in qs:
                q = q & item
            receipts = Receipt.objects.filter(q)
        return receipts


class SendReceiptEmail(StaffRequiredMixin, View):
    def get(self, request, receipt_id, *args, **kwargs):
        receipt = Receipt.objects.get(pk=receipt_id)
        services = ReceiptService.objects.filter(receipt_id=receipt_id)
        context = {
            'services': services,
            'receipt': receipt,
        }
        html_template = get_template('cabinet/invoice/invoice.html')
        html = html_template.render(context)
        base_url = request.build_absolute_uri()

        send_receipt.delay(html_to_pdf=html, base_url=base_url,
                           to=receipt.flat.flat_owner.user.email)
        url = f"/admin/receipt/detail/{receipt_id}"
        return HttpResponseRedirect(url)


class UpdateReceipt(StaffRequiredMixin, UpdateView):
    model = Receipt
    template_name = 'admin_panel/update_receipt.html'
    form_class = ReceiptForm
    success_url = reverse_lazy('receipts')

    def get(self, request, *args, **kwargs):
        receipt = Receipt.objects.get(id=self.kwargs['pk'])
        data = {
            'indications': Indication.objects.order_by('date_published').filter(flat_id=receipt.flat.id),
            'form': ReceiptForm(instance=receipt,
                                initial={'house': receipt.flat.house, 'section': receipt.flat.section}),
            'service_formset': ReceiptServiceFormset(prefix='service', queryset=ReceiptService.objects.filter(
                receipt_id=self.kwargs['pk'])),
            'services': Service.objects.all(),
            'measures': Measure.objects.all(),
        }
        return render(request, 'admin_panel/update_receipt.html', context=data)

    def post(self, request, pk, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST, instance=Receipt.objects.get(id=pk))

        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save()
            for instance in instances:
                instance.receipt_id = obj.id
                instance.save()

            if hasattr(obj.flat, 'personal_account'):
                if obj.flat.personal_account is not None \
                        and obj.flat.personal_account != '' \
                        and obj.is_complete is True:
                    personal_account = PersonalAccount.objects.get(pk=obj.flat.personal_account.id)
                    plus_total = sum(
                        Paybox.objects.filter(personal_account=personal_account,
                                              is_complete=True).values_list('total', flat=True))
                    minus_total = sum(
                        Receipt.objects.filter(flat__personal_account=personal_account,
                                               is_complete=True).values_list('total_price', flat=True))

                    personal_account.balance = plus_total - minus_total
                    personal_account.save()
            return redirect('receipts')
        else:
            data = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, 'admin_panel/update_receipt.html', context=data)


class ReceiptPrint(StaffRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        receipt = Receipt.objects.get(pk=pk)
        rows = ReceiptExcelDoc.objects.all().order_by('id')
        data = {
            'rows': rows,
            'receipt': receipt
        }
        return render(request, 'admin_panel/receipt_printing.html', context=data)


def copy_row(ws, source_row_num, target_row_num):
    # Get the source row
    source_row = ws[source_row_num]

    # Copy the source row to a new row number
    for cell in source_row:
        if cell.coordinate in ws.merged_cells:
            for merged_cell in list(ws.merged_cells.ranges):
                if cell.coordinate in merged_cell:
                    merged_range = ws.cell(row=target_row_num,
                                           column=merged_cell.min_col).coordinate + ":" + \
                                   ws.cell(row=target_row_num,
                                           column=merged_cell.max_col).coordinate
                    ws.merge_cells(merged_range)
                    ws.cell(row=target_row_num, column=merged_cell.min_col, value=cell.value)
                    ws.cell(row=target_row_num, column=merged_cell.min_col).alignment = Alignment(horizontal='left',
                                                                                                  vertical='center')  # выровнять по центр
        else:
            ws.cell(row=target_row_num, column=cell.col_idx, value=cell.value)

    target_row = ws[target_row_num]  # Get the destination row

    # Copy styles from source row to destination row
    for source_cell, dest_cell in zip(source_row, target_row):
        dest_cell.font = copy(source_cell.font)  # Copy font style
        dest_cell.fill = copy(source_cell.fill)  # Copy fill style
        dest_cell.border = copy(source_cell.border)  # Copy border style


class ReceiptDownloadExcel(StaffRequiredMixin, View):
    def post(self, request, excel_id, receipt_id, *args, **kwargs):
        receipt = Receipt.objects.get(pk=receipt_id)
        services = ReceiptService.objects.filter(receipt_id=receipt_id)
        services_count = services.count()
        doc = ReceiptExcelDoc.objects.get(pk=excel_id)

        wb = load_workbook(doc.file)
        ws = wb.active  # это лист в excel
        if services_count > 0:
            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=50):
                for cell in row:
                    if cell.value == 'total':
                        final_row = ws[cell.row]
                        ws.delete_rows(cell.row)

            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=50):
                for cell in row:
                    match cell.value:
                        case 'service':
                            for i, obj in enumerate(services):
                                ws.cell(row=cell.row + i, column=cell.column, value=str(obj.service.title))
                                copy_row(ws, cell.row, cell.row + i)
                                ws.cell(row=cell.row + i, column=cell.column, value=str(obj.service.title))
                            else:
                                for cell in final_row:
                                    if cell.coordinate in ws.merged_cells:
                                        for merged_cell in list(ws.merged_cells.ranges):
                                            if cell.coordinate in merged_cell:
                                                merged_range = ws.cell(row=cell.row + i,
                                                                       column=merged_cell.min_col).coordinate + ":" + \
                                                               ws.cell(row=cell.row + i,
                                                                       column=merged_cell.max_col).coordinate
                                                ws.merge_cells(merged_range)
                                                ws.cell(row=cell.row + i, column=merged_cell.min_col, value=cell.value)
                                                ws.cell(row=cell.row + i,
                                                        column=merged_cell.min_col).alignment = Alignment(
                                                    horizontal='right', vertical='center')  # выровнять по центр
                                    else:
                                        ws.cell(row=cell.row + i, column=cell.col_idx, value=cell.value)
                                for source_cell, dest_cell in zip(final_row, ws[cell.row + i]):
                                    dest_cell.font = copy(source_cell.font)  # Copy font style
                                    dest_cell.fill = copy(source_cell.fill)  # Copy fill style
                                    dest_cell.border = copy(source_cell.border)  # Copy border style
                        case 'tariff':
                            for i, obj in enumerate(services):
                                ws.cell(row=cell.row + i, column=cell.column, value=str(obj.receipt.tariff.title))
                        case 'measure':
                            for i, obj in enumerate(services):
                                ws.cell(row=cell.row + i, column=cell.column, value=str(obj.measure.title))
                        case 'totalServicePrice':
                            for i, obj in enumerate(services):
                                ws.cell(row=cell.row + i, column=cell.column,
                                        value=str(obj.unit_price * obj.consumption))
        else:
            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=50):
                for cell in row:
                    match cell.value:
                        case 'service':
                            ws.delete_rows(cell.row)

        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=50):
            for cell in row:
                match cell.value:
                    case 'personal_accountNumber':
                        if hasattr(receipt.flat, 'personal_account'):
                            if receipt.flat.personal_account is None or receipt.flat.personal_account == "":
                                cell.value = ''
                            else:
                                cell.value = receipt.flat.personal_account.number
                        else:
                            cell.value = ''
                    case 'personalManager':
                        if receipt.flat.flat_owner is None:
                            cell.value = ''
                        else:
                            cell.value = str(receipt.flat.flat_owner)
                            cell.alignment = Alignment(horizontal='center', vertical='center')  # выровнять по центру
                    case 'receiptNumber':
                        cell.value = receipt.number
                    case 'receiptStartDate':
                        cell.value = str(receipt.start_date)
                    case 'totalPrice':
                        cell.value = receipt.total_price
                    case 'flatOwner':
                        if receipt.flat.flat_owner is None:
                            cell.value = ''
                        else:
                            cell.value = str(receipt.flat.flat_owner)
                    case 'personalAccountBalance':
                        if hasattr(receipt.flat, 'personal_account'):
                            if receipt.flat.personal_account is None or receipt.flat.personal_account == "":
                                cell.value = ''
                            else:
                                cell.value = receipt.flat.personal_account.balance
                        else:
                            cell.value = ''
                    case 'receiptDatePublished':
                        cell.value = receipt.date_published
                        cell.alignment = Alignment(horizontal='center', vertical='center')  # выровнять по центр
                    case 'receiptMonthPublished':
                        month_dict = {
                            'January': 'Январь',
                            'February': 'Февраль',
                            'March': 'Март',
                            'April': 'Апрель',
                            'May': 'Май',
                            'June': 'Июнь',
                            'July': 'Июль',
                            'August': 'Август',
                            'September': 'Сентябрь',
                            'October': 'Октябрь',
                            'November': 'Ноябрь',
                            'December': 'Декабрь'
                        }
                        cell.value = month_dict[str(receipt.date_published.strftime('%B'))]
                        cell.alignment = Alignment(horizontal='center', vertical='center')  # выровнять по центр
                    case 'total':
                        cell.value = receipt.total_price
        ws.title = "Квитанция"  # это название листа в excel
        wb.save('media/receipt/result/report.xlsx')

        data = {
        }
        return HttpResponse(json.dumps(data), content_type='application/json')


class ReceiptPrintingSettings(StaffRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        rows = ReceiptExcelDoc.objects.all().order_by('id')
        form = ReceiptExcelDocForm()
        data = {
            'rows': rows,
            'form': form,
        }
        return render(request, 'admin_panel/receipt_printing_settings.html', context=data)

    def post(self, request, *args, **kwargs):
        form = ReceiptExcelDocForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
        return redirect('receipt_print_settings')


class ReceiptPrintingSettingsDefault(StaffRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        for row in ReceiptExcelDoc.objects.all():
            row.by_default = False
            row.save()
        row = ReceiptExcelDoc.objects.get(pk=pk)
        row.by_default = True
        row.save()
        return redirect('receipt_print_settings')


class ReceiptPrintingSettingsDelete(StaffRequiredMixin, DeleteView):
    success_url = reverse_lazy('receipt_print_settings')
    queryset = ReceiptExcelDoc.objects.all()


class CopyReceipt(StaffRequiredMixin, FormView):
    def get(self, request, pk, *args, **kwargs):
        copy = Receipt.objects.get(pk=pk)
        receipt_form = ReceiptForm(instance=copy, initial={'house': copy.flat.house, 'section': copy.flat.section})

        service_formset = ReceiptServiceFormset(queryset=ReceiptService.objects.filter(receipt_id=pk),
                                                prefix='service')
        data = {
            'form': receipt_form,
            'service_formset': service_formset,
            "indications": Indication.objects.order_by('date_published').filter(flat_id=copy.flat.id),
            'measures': Measure.objects.all(),
            'services': Service.objects.all(),
        }
        return render(request, 'admin_panel/get_receipt_form.html', context=data)

    def post(self, request, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST)
        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save(commit=False)
            for instance in instances:
                instance.receipt_id = obj.id
                instance.save()
            return redirect('receipts')
        else:
            data = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, 'admin_panel/get_receipt_form.html', context=data)


class ReceiptDetail(StaffRequiredMixin, DetailView):
    model = Receipt
    template_name = 'admin_panel/read_receipt.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        receipt = Receipt.objects.get(pk=self.kwargs['pk'])
        receipt_services = ReceiptService.objects.filter(receipt_id=self.kwargs['pk'])
        context['receipt'] = receipt
        context['receipt_services'] = receipt_services
        return context


class DeleteReceipt(StaffRequiredMixin, DeleteView):
    success_url = reverse_lazy('receipts')
    queryset = Receipt.objects.all()

    def post(self, request, *args, **kwargs):
        receipt = Receipt.objects.get(pk=self.kwargs['pk'])
        if hasattr(receipt.flat, 'personal_account'):
            if receipt.flat.personal_account is not None \
                    and receipt.flat.personal_account != '' \
                    and receipt.is_complete is True:
                personal_account = PersonalAccount.objects.get(pk=receipt.flat.personal_account.id)
                receipt.delete()
                plus_total = sum(
                    Paybox.objects.filter(personal_account=personal_account,
                                          is_complete=True).values_list('total', flat=True))
                minus_total = sum(
                    Receipt.objects.filter(flat__personal_account=personal_account,
                                           is_complete=True).values_list('total_price', flat=True))

                personal_account.balance = plus_total - minus_total
                personal_account.save()
        else:
            receipt.delete()
        return redirect('receipts')


class FlatListView(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/flats.html'
    context_object_name = 'flats'
    queryset = Flat.objects.all()
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = FlatsFilterForm()
        return context


class FlatFilteredList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/flats.html'
    context_object_name = 'flats'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = FlatsFilterForm(initial=self.request.GET)
        return context

    def get_queryset(self):
        flats = Flat.objects.all()
        form_filter = FlatsFilterForm(self.request.GET)
        qs = []
        if form_filter.is_valid():
            if form_filter.cleaned_data['number']:
                qs.append(Q(number=form_filter.cleaned_data['number']))
            if form_filter.cleaned_data['house']:
                qs.append(Q(house_id=form_filter.cleaned_data['house'].id))
            if form_filter.cleaned_data['section']:
                qs.append(Q(section_id=form_filter.cleaned_data['section'].id))
            if form_filter.cleaned_data['floor']:
                qs.append(Q(floor_id=form_filter.cleaned_data['floor'].id))
            if form_filter.cleaned_data['flat_owner']:
                full_name = str(form_filter.cleaned_data['flat_owner']).split(' ')
                qs.append(Q(
                    Q(flat_owner__patronymic__icontains=full_name[0]) |
                    Q(flat_owner__user__first_name__icontains=full_name[1]) |
                    Q(flat_owner__user__last_name__icontains=full_name[2])
                ))
            if form_filter.cleaned_data['have_debts']:
                if form_filter.cleaned_data['have_debts'] == 'no':
                    qs.append(Q(personal_account__balance__gte=0))
                elif form_filter.cleaned_data['have_debts'] == 'yes':
                    qs.append(Q(personal_account__balance__lt=0))
            q = Q()
            for item in qs:
                q = q & item
            flats = Flat.objects.filter(q)
        return flats


class CreateFlatView(StaffRequiredMixin, CreateView):
    model = Flat
    template_name = 'admin_panel/get_flat_form.html'
    form_class = FlatForm
    success_url = reverse_lazy('flats')


class FlatDetail(StaffRequiredMixin, DetailView):
    model = Flat
    template_name = 'admin_panel/read_flat.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        flat = Flat.objects.get(pk=self.kwargs['pk'])
        context['flat'] = flat
        return context


class UpdateFlatView(StaffRequiredMixin, UpdateView):
    model = Flat
    template_name = 'admin_panel/update_flat.html'
    form_class = FlatForm
    success_url = reverse_lazy('flats')


class DeleteFlatView(StaffRequiredMixin, DeleteView):
    success_url = reverse_lazy('flats')
    queryset = Flat.objects.all()


class FlatAcceptPayment(CreatePaybox):
    def get(self, request, pk, *args, **kwargs):
        initial = {}
        flat = Flat.objects.get(pk=pk)
        if flat.personal_account is not None and flat.flat_owner is not None:
            initial['flat_owner'] = flat.flat_owner
            initial['personal_account'] = flat.personal_account

        form = PayboxForm(initial=initial)
        form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        form.fields['date_published'].initial = timezone.now().date()
        form.fields['user'].initial = request.user.personal.id
        form.fields['is_complete'].initial = True

        data = {
            'income': 'plus',
            'form': form,
        }
        return render(request, 'admin_panel/get_paybox_form.html', context=data)


class FlatAcceptReceipt(StaffRequiredMixin, FormView):
    def get(self, request, pk, *args, **kwargs):
        flat = Flat.objects.get(pk=pk)
        init_data = {
            'house': flat.house,
            'section': flat.section,
            'flat': flat,
            'tariff': flat.tariff,
        }
        if hasattr(flat, 'personal_account'):
            if flat.personal_account is not None and flat.personal_account != '':
                init_data['personal_account'] = flat.personal_account

        receipt_form = ReceiptForm(
            initial=init_data
        )

        service_formset = ReceiptServiceFormset(queryset=ReceiptService.objects.none(),
                                                prefix='service')
        data = {
            'form': receipt_form,
            'service_formset': service_formset,
            "indications": Indication.objects.order_by('date_published').filter(flat_id=flat.id),
            'measures': Measure.objects.all(),
            'services': Service.objects.all(),
        }
        return render(request, 'admin_panel/get_receipt_form.html', context=data)

    def post(self, request, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST)
        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save(commit=False)
            for instance in instances:
                instance.receipt_id = obj.id
                instance.save()
            return redirect('receipts')
        else:
            data = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, 'admin_panel/get_receipt_form.html', context=data)


class FlatReceiptList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/receipts.html'
    context_object_name = 'receipts'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        flat = Flat.objects.get(pk=self.kwargs['pk'])
        context['filter_form'] = ReceiptFilterForm(initial={'flat': flat.number})
        return context

    def get_queryset(self):
        flat = Flat.objects.get(pk=self.kwargs['pk'])
        queryset = Receipt.objects.filter(flat_id=flat.id)
        return queryset


class FlatPayboxList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/paybox.html'
    context_object_name = 'paybox'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        flat = Flat.objects.get(pk=self.kwargs['pk'])
        context['filter_form'] = PayboxFilterForm(initial={'personal_account': flat.personal_account})

        if flat.personal_account is not None:
            total_plus = sum(Paybox.objects.filter(debit_credit='plus',
                                                   personal_account=flat.personal_account,
                                                   is_complete=True).values_list('total', flat=True))
        else:
            total_plus = 0
        context['total_plus'] = total_plus
        return context

    def get_queryset(self):
        flat = Flat.objects.get(pk=self.kwargs['pk'])
        if flat.personal_account is not None:
            queryset = Paybox.objects.filter(personal_account=flat.personal_account)
        else:
            queryset = Paybox.objects.none()

        return queryset


class PersonalAccountListView(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/personal_accounts.html'
    context_object_name = 'personal_accounts'
    paginate_by = 5

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = PersonalAccountsFilterForm()
        context['personal_accounts_count'] = PersonalAccount.objects.all().count()
        balances(context)
        return context

    def get_queryset(self):
        personal_accounts = PersonalAccount.objects.all()

        wb = Workbook()
        ws = wb.active  # это лист в excel
        ws.append(
            [
                'Лицевой счет',
                'Cтатус',
                'Дом',
                'Секция',
                'Квартира',
                'Владелец',
                'Остаток',
            ]
        )
        # Define a font style that is bold
        bold_font = Font(bold=True)
        # Apply the bold font style to the cells in the row
        for cell in ws[1]:
            cell.font = bold_font

        for obj in personal_accounts:
            if obj.house is None:
                house = f''
            else:
                house = f'{obj.house}'

            if obj.section is None:
                section = f''
            else:
                section = f'{obj.section}'

            if obj.flat is None:
                flat = f''
            else:
                flat = f'{obj.flat}'

            if obj.flat is not None:
                if obj.flat.flat_owner is not None:
                    flat_owner = f'{obj.flat.flat_owner}'
                else:
                    flat_owner = f''
            else:
                flat_owner = f''
            ws.append([
                f'{obj.number}',
                f'{obj.get_status_display()}',
                f'{house}',
                f'{section}',
                f'{flat}',
                f'{flat_owner}',
                f'{obj.balance}',
            ])

        ws.title = "Выписка"  # это название листа в excel
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 20

        wb.save('media/personal_account/info.xlsx')

        # Get full path to workbook
        wb.close()

        return personal_accounts


class PersonalAccountFilteredList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/personal_accounts.html'
    context_object_name = 'personal_accounts'
    paginate_by = 5

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = PersonalAccountsFilterForm(initial=self.request.GET)
        balances(context)
        return context

    def get_queryset(self):
        personal_accounts = PersonalAccount.objects.all()
        form_filter = PersonalAccountsFilterForm(self.request.GET)
        qs = []
        if form_filter.is_valid():
            if form_filter.cleaned_data['number']:
                qs.append(Q(number__icontains=form_filter.cleaned_data['number']))
            if form_filter.cleaned_data['status']:
                qs.append(Q(status=form_filter.cleaned_data['status']))
            if form_filter.cleaned_data['flat']:
                qs.append(Q(flat__number=form_filter.cleaned_data['flat']))
            if form_filter.cleaned_data['house']:
                qs.append(Q(house_id=form_filter.cleaned_data['house'].id))
            if form_filter.cleaned_data['section']:
                qs.append(Q(section_id=form_filter.cleaned_data['section'].id))
            if form_filter.cleaned_data['flat_owner']:
                full_name = str(form_filter.cleaned_data['flat_owner']).split()
                qs.append(Q(
                    Q(flat__flat_owner__patronymic=full_name[2]) &
                    Q(flat__flat_owner__user__first_name=full_name[1]) &
                    Q(flat__flat_owner__user__last_name=full_name[0])
                ))
            if form_filter.cleaned_data['have_debts']:
                if form_filter.cleaned_data['have_debts'] == 'no':
                    qs.append(Q(balance__gte=0))
                elif form_filter.cleaned_data['have_debts'] == 'yes':
                    qs.append(Q(balance__lt=0))
            q = Q()
            for item in qs:
                q = q & item
            personal_accounts = PersonalAccount.objects.filter(q)
        return personal_accounts


class PersonalAccountDetail(StaffRequiredMixin, DetailView):
    model = PersonalAccount
    template_name = 'admin_panel/read_personal_account.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        personal_account = PersonalAccount.objects.get(pk=self.kwargs['pk'])
        context['personal_account'] = personal_account
        return context


class CreatePersonalAccount(StaffRequiredMixin, CreateView):
    model = PersonalAccount
    template_name = 'admin_panel/get_personal_account_form.html'
    form_class = PersonalAccountForm
    success_url = reverse_lazy('personal_accounts')


class UpdatePersonalAccount(StaffRequiredMixin, UpdateView):
    model = PersonalAccount
    template_name = 'admin_panel/update_personal_account.html'
    form_class = PersonalAccountForm
    success_url = reverse_lazy('personal_accounts')


class DeletePersonalAccount(StaffRequiredMixin, DeleteView):
    success_url = reverse_lazy('personal_accounts')
    queryset = PersonalAccount.objects.all()


class PersonalAccountAcceptPayment(CreatePaybox):
    def get(self, request, pk, *args, **kwargs):
        initial = {}
        personal_account = PersonalAccount.objects.get(pk=pk)
        initial['personal_account'] = personal_account
        if personal_account.flat is not None and personal_account.flat.flat_owner is not None:
            initial['flat_owner'] = personal_account.flat.flat_owner
        form = PayboxForm(initial=initial)
        form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        form.fields['date_published'].initial = timezone.now().date()
        form.fields['user'].initial = request.user.personal.id
        form.fields['is_complete'].initial = True

        data = {
            'income': 'plus',
            'form': form,
        }
        return render(request, 'admin_panel/get_paybox_form.html', context=data)


class PersonalAccountAcceptReceipt(StaffRequiredMixin, FormView):
    def get(self, request, pk, *args, **kwargs):
        personal_account = PersonalAccount.objects.get(pk=pk)

        receipt_form = ReceiptForm(
            initial={
                'house': personal_account.flat.house,
                'section': personal_account.flat.section,
                'flat': personal_account.flat,
                'tariff': personal_account.flat.tariff,
                'personal_account': personal_account,
            }
        )

        service_formset = ReceiptServiceFormset(queryset=ReceiptService.objects.none(),
                                                prefix='service')
        data = {
            'form': receipt_form,
            'service_formset': service_formset,
            "indications": Indication.objects.order_by('date_published').filter(flat_id=personal_account.flat.id),
            'measures': Measure.objects.all(),
            'services': Service.objects.all(),
        }
        return render(request, 'admin_panel/get_receipt_form.html', context=data)

    def post(self, request, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST)
        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save(commit=False)
            for instance in instances:
                instance.receipt_id = obj.id
                instance.save()
            return redirect('receipts')
        else:
            data = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, 'admin_panel/get_receipt_form.html', context=data)


class PersonalAccountReceiptList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/receipts.html'
    context_object_name = 'receipts'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        personal_account = PersonalAccount.objects.get(pk=self.kwargs['pk'])
        context['filter_form'] = ReceiptFilterForm(initial={'flat': personal_account.flat.number})
        return context

    def get_queryset(self):
        personal_account = PersonalAccount.objects.get(pk=self.kwargs['pk'])
        if personal_account.flat is not None:
            queryset = Receipt.objects.filter(flat_id=personal_account.flat.id)
        else:
            queryset = Receipt.objects.none()
        return queryset


class PersonalAccountPayboxList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/paybox.html'
    context_object_name = 'paybox'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        personal_account = PersonalAccount.objects.get(pk=self.kwargs['pk'])
        context['filter_form'] = PayboxFilterForm(initial={'personal_account': personal_account.number})

        total_plus = sum(Paybox.objects.filter(debit_credit='plus',
                                               personal_account_id=self.kwargs['pk'],
                                               is_complete=True).values_list('total', flat=True))

        context['total_plus'] = total_plus
        return context

    def get_queryset(self):
        personal_account = PersonalAccount.objects.get(pk=self.kwargs['pk'])
        queryset = Paybox.objects.filter(personal_account=personal_account)
        return queryset


class ClientListView(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/clients.html'
    context_object_name = 'clients'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ClientsFilterForm()
        return context

    def get_queryset(self):
        clients = FlatOwner.objects.all()
        return clients


class ClientFilteredListView(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/clients.html'
    context_object_name = 'clients'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ClientsFilterForm(initial=self.request.GET)
        return context

    def get_queryset(self):
        clients = FlatOwner.objects.all()
        form_filter = ClientsFilterForm(self.request.GET)
        qs = []
        if form_filter.is_valid():

            if form_filter.cleaned_data['ID']:
                qs.append(Q(ID__icontains=form_filter.cleaned_data['ID']))
            if form_filter.cleaned_data['full_name']:
                full_name = form_filter.cleaned_data['full_name'].split(' ')
                qs.append(Q(
                    Q(patronymic__icontains=form_filter.cleaned_data['full_name']) |
                    Q(patronymic__in=full_name) |
                    Q(user__first_name__in=full_name) |
                    Q(user__last_name__in=full_name) |
                    Q(user__first_name__icontains=form_filter.cleaned_data['full_name']) |
                    Q(user__last_name__icontains=form_filter.cleaned_data['full_name'])
                ))

            if form_filter.cleaned_data['phone']:
                qs.append(Q(user__phone__icontains=form_filter.cleaned_data['phone']))
            if form_filter.cleaned_data['email']:
                qs.append(Q(user__email=form_filter.cleaned_data['email']))
            if form_filter.cleaned_data['house']:
                qs.append(Q(flat__house_id=form_filter.cleaned_data['house'].id))
            if form_filter.cleaned_data['flat']:
                qs.append(Q(flat__number=form_filter.cleaned_data['flat']))
            if form_filter.cleaned_data['status']:
                qs.append(Q(user__status=form_filter.cleaned_data['status']))
            if form_filter.cleaned_data['date_added']:
                date_str = form_filter.cleaned_data['date_added'].split('/')
                date_str.reverse()
                result = "-".join(date_str)
                qs.append(Q(user__date_joined=result))

            q = Q()
            for item in qs:
                q = q & item
            clients = FlatOwner.objects.filter(q)
        return clients


class ClientDetail(StaffRequiredMixin, DetailView):
    model = FlatOwner
    template_name = 'admin_panel/read_client.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        client = FlatOwner.objects.prefetch_related("flat_set").get(pk=self.kwargs['pk'])

        context['client'] = client
        context['flats'] = client.flat_set.all()

        return context


from admin_panel.tasks import send_invitation


class SendInvitation(StaffRequiredMixin, FormView):
    template_name = 'admin_panel/send_invitation.html'
    form_class = InvitationForm
    success_url = reverse_lazy('send_invitation')

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        data = {
            'form': form
        }
        return render(self.request, 'admin_panel/send_invitation.html', context=data)

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            send_invitation.delay(to=form.cleaned_data['email'])
            return redirect('send_invitation')
        else:
            data = {
                'form': form
            }
            return render(self.request, 'admin_panel/send_invitation.html', context=data)


class ClientSignUpView(StaffRequiredMixin, CreateView):
    model = CustomUser
    form_class = ClientSignUpForm
    template_name = 'admin_panel/get_client_form.html'

    def get_context_data(self, **kwargs):
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        form.save()
        return redirect('clients')


class UpdateClientView(StaffRequiredMixin, UpdateView):
    model = CustomUser
    form_class = ClientUpdateForm
    template_name = 'admin_panel/update_client.html'
    success_url = reverse_lazy('clients')
    queryset = CustomUser.objects.all()

    def get(self, request, pk, *args, **kwargs):
        user = CustomUser.objects.get(pk=pk)
        owner = FlatOwner.objects.get(user=user)
        form = ClientUpdateForm(instance=user,
                                initial={'birthday': owner.birthday, 'viber': owner.viber, 'telegram': owner.telegram,
                                         'patronymic': owner.patronymic, 'ID': owner.ID,
                                         'bio': owner.bio, })
        data = {
            'form': form,
        }
        return render(request, 'admin_panel/update_client.html', context=data)


class DeleteClientView(StaffRequiredMixin, DeleteView):
    success_url = reverse_lazy('clients')
    queryset = CustomUser.objects.all()


class HouseListView(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/houses.html'
    context_object_name = 'houses'
    queryset = House.objects.all()
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = HouseFilterForm()
        return context


class HouseFilteredList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/houses.html'
    context_object_name = 'houses'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = HouseFilterForm(initial=self.request.GET)
        return context

    def get_queryset(self):
        houses = House.objects.all()
        form_filter = HouseFilterForm(self.request.GET)
        qs = []
        if form_filter.is_valid():
            if form_filter.cleaned_data['title']:
                qs.append(Q(title__icontains=form_filter.cleaned_data['title']))
            if form_filter.cleaned_data['address']:
                qs.append(Q(address__icontains=form_filter.cleaned_data['address']))
            q = Q()
            for item in qs:
                q = q & item
            houses = House.objects.filter(q)
        return houses


class HouseDetail(StaffRequiredMixin, DetailView):
    model = House
    template_name = 'admin_panel/read_house.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        house = House.objects.prefetch_related('gallery__photo_set', 'houseuser_set').get(pk=self.kwargs['pk'])
        photos = house.gallery.photo_set.all()
        users = house.houseuser_set.all()
        context['house'] = house
        context['photos'] = photos
        context['users'] = users
        return context


class CreateHouseView(StaffRequiredMixin, FormView):
    def get(self, request, *args, **kwargs):
        house_form = HouseForm()
        personals = Personal.objects.all()
        photo_formset = HousePhotoFormset(queryset=Photo.objects.none(), prefix='gallery')
        section_formset = SectionFormset(queryset=Section.objects.none(), prefix='section')
        floor_formset = FloorFormset(queryset=Floor.objects.none(), prefix='floor')
        house_user_formset = HouseUserFormset(queryset=HouseUser.objects.none(), prefix='personal')
        data = {
            'house_form': house_form,
            'personals': personals,
            'photo_formset': photo_formset,
            'section_formset': section_formset,
            'floor_formset': floor_formset,
            'house_user_formset': house_user_formset,
        }
        return render(request, 'admin_panel/get_house_form.html', context=data)

    def post(self, request, *args, **kwargs):

        house_form = HouseForm(request.POST, request.FILES)
        section_formset = SectionFormset(request.POST, request.FILES, prefix='section')
        floor_formset = FloorFormset(request.POST, request.FILES, prefix='floor')
        photo_formset = HousePhotoFormset(request.POST, request.FILES, prefix='gallery')
        house_user_formset = HouseUserFormset(request.POST, request.FILES, prefix='personal')

        if section_formset.is_valid() and floor_formset.is_valid() and house_form.is_valid() and photo_formset.is_valid() and house_user_formset.is_valid():
            house_obj = house_form.save()
            house_user_instances = house_user_formset.save(commit=False)
            for instance in house_user_instances:
                instance.house_id = house_obj.id
                instance.save()
            gallery = Gallery.objects.create()
            house_obj.gallery_id = gallery.id
            for photo_form in photo_formset:
                instance = photo_form.save(commit=False)
                instance.gallery_id = gallery.id
                instance.save()
            section_instances = section_formset.save(commit=False)
            floor_instances = floor_formset.save(commit=False)
            for instance in section_instances:
                instance.house_id = house_obj.id

                instance.save()
            for instance in floor_instances:
                instance.house_id = house_obj.id
                instance.save()
            house_obj.save()

        else:
            data = {
                'house_form': house_form,
                'section_formset': section_formset,
                'floor_formset': floor_formset,
                'photo_formset': photo_formset,
                'house_user_formset': house_user_formset,
            }
            return render(request, 'admin_panel/get_house_form.html', context=data)
        return redirect('houses')


class UpdateHouseView(StaffRequiredMixin, FormView):
    def get(self, request, pk, *args, **kwargs):
        personals = Personal.objects.all()
        house = House.objects.get(pk=pk)
        house_form = HouseForm(instance=house)
        photo_formset = PhotoFormset(queryset=Photo.objects.filter(gallery_id=house.gallery_id), prefix='gallery')
        section_formset = SectionFormset(queryset=Section.objects.filter(house=house), prefix='section')
        floor_formset = FloorFormset(queryset=Floor.objects.filter(house=house), prefix='floor')
        house_user_formset = HouseUserFormset(queryset=HouseUser.objects.filter(house_id=house.id), prefix='personal')
        data = {
            'house_form': house_form,
            'personals': personals,
            'house': house,
            'photo_formset': photo_formset,
            'section_formset': section_formset,
            'floor_formset': floor_formset,
            'house_user_formset': house_user_formset,
        }
        return render(request, 'admin_panel/update_house.html', context=data)

    def post(self, request, pk, *args, **kwargs):
        house = House.objects.get(pk=pk)
        house_form = HouseForm(request.POST, request.FILES, instance=house)
        section_formset = SectionFormset(request.POST, request.FILES, prefix='section')
        floor_formset = FloorFormset(request.POST, request.FILES, prefix='floor')
        photo_formset = PhotoFormset(request.POST, request.FILES, prefix='gallery')
        house_user_formset = HouseUserFormset(request.POST, request.FILES, prefix='personal')

        if section_formset.is_valid() and floor_formset.is_valid() and house_form.is_valid() and photo_formset.is_valid() and house_user_formset.is_valid():
            house_obj = house_form.save()
            house_user_instances = house_user_formset.save(commit=False)
            for instance in house_user_instances:
                instance.house_id = house_obj.id
                instance.save()
            photo_instances = photo_formset.save(commit=False)
            gallery = Gallery.objects.get(house=house)
            for photo_instance in photo_instances:
                photo_instance.gallery_id = gallery.id
                photo_instance.save()

            section_instances = section_formset.save(commit=False)
            floor_instances = floor_formset.save(commit=False)
            for instance in section_instances:
                instance.house_id = house_obj.id
                instance.save()
            for instance in floor_instances:
                instance.house_id = house_obj.id
                instance.save()
            house_obj.save()

        else:
            data = {
                'house_form': house_form,
                'section_formset': section_formset,
                'floor_formset': floor_formset,
                'photo_formset': photo_formset,
                'house_user_formset': house_user_formset,
            }
            return render(request, 'admin_panel/get_house_form.html', context=data)
        return redirect('houses')


class GetRoleView(StaffRequiredMixin, View):
    def get(self, request, pk):
        personal = Personal.objects.get(pk=pk)
        data = {
            'role': personal.get_role_display(),
        }

        return HttpResponse(json.dumps(data), content_type='application/json')


class GetSectionInfoView(StaffRequiredMixin, View):
    def get(self, request, pk):
        section = Section.objects.prefetch_related('flat_set').get(pk=pk)
        flats = serializers.serialize('json', section.flat_set.all())
        data = {
            "flats": flats,
        }
        return JsonResponse(data, safe=False)


class GetFlatInfoView(StaffRequiredMixin, View):
    def get(self, request, pk):
        data = {}
        flat = Flat.objects.get(pk=pk)
        try:
            flat_owner_obj = FlatOwner.objects.get(flat=flat)
            flat_owner = serializers.serialize('json', [flat_owner_obj])
            data['flat_owner'] = flat_owner
            user = CustomUser.objects.get(client=flat_owner_obj)
            user = serializers.serialize('json', [user])
            data['user'] = user
        except ObjectDoesNotExist:
            pass
        try:
            if flat.personal_account is not None:
                personal_account = flat.personal_account
                personal_account = serializers.serialize('json', [personal_account])
                data['personal_account'] = personal_account
        except ObjectDoesNotExist:
            pass
        try:
            if flat.tariff is not None:
                tariff = TariffSystem.objects.get(pk=flat.tariff.pk)
                tariff = serializers.serialize('json', [tariff])
                data['tariff'] = tariff
        except ObjectDoesNotExist:
            pass
        return JsonResponse(data, safe=False)


class GetHouseInfoView(StaffRequiredMixin, View):
    def get(self, request, pk):
        house = House.objects.prefetch_related('section_set', 'floor_set', 'flat_set').get(pk=pk)
        sections = serializers.serialize('json', house.section_set.all())
        floors = serializers.serialize('json', house.floor_set.all())
        flats = serializers.serialize('json', house.flat_set.all())
        house = serializers.serialize('json', [house])
        data = {
            "house": house,
            "sections": sections,
            "floors": floors,
            "flats": flats,
        }
        return JsonResponse(data, safe=False)


class GetFlatsForMailbox(StaffRequiredMixin, View):
    def get(self, request, section_id, floor_id):
        flats = Flat.objects.all()
        if self.kwargs['section_id'] != "None":
            flats = flats.filter(section_id=self.kwargs['section_id'])
        if self.kwargs['floor_id'] != "None":
            flats = flats.filter(floor_id=self.kwargs['floor_id'])

        flats = serializers.serialize('json', flats)
        data = {
            "flats": flats,
        }
        return JsonResponse(data, safe=False)


class GetTariffInfoView(StaffRequiredMixin, View):
    def get(self, request, pk):
        tariff = TariffSystem.objects.prefetch_related('tariffservice_set').get(pk=pk)
        tariff_services = serializers.serialize('json', tariff.tariffservice_set.all())
        data = {
            "tariff_services": tariff_services,
        }
        return JsonResponse(data, safe=False)


class GetServiceInfoView(StaffRequiredMixin, View):
    def get(self, request, pk):
        service = Service.objects.get(pk=pk)
        service = serializers.serialize('json', [service])
        data = {
            "service": service,
        }
        return JsonResponse(data, safe=False)


class GetIndicationInfoView(StaffRequiredMixin, View):
    def get(self, request, flat_id, service_id):
        indication = Indication.objects.filter(flat_id=flat_id, service_id=service_id)
        if indication.count() != 0:
            indication = serializers.serialize('json', indication)
            data = {
                "indication": indication,
            }
        else:
            return JsonResponse({}, safe=False)

        return JsonResponse(data, safe=False)


class GetFlatOwnerInfo(StaffRequiredMixin, View):
    def get(self, request, pk):
        flat_owner = FlatOwner.objects.prefetch_related('flat_set').get(pk=pk)
        flats = flat_owner.flat_set.select_related('house').all()
        personal_accounts = PersonalAccount.objects.filter(flat__in=flats.values_list("id", flat=True))
        flats = serializers.serialize('json', flats)
        personal_accounts = serializers.serialize('json', personal_accounts)
        data = {
            "flats": flats,
            "personal_accounts": personal_accounts,
        }
        return JsonResponse(data, safe=False)


class GetAllFlats(StaffRequiredMixin, View):
    def get(self, request):
        flats = Flat.objects.all()
        flats = serializers.serialize('json', flats)
        all_personal_accounts = PersonalAccount.objects.all()
        all_personal_accounts = serializers.serialize('json', all_personal_accounts)

        data = {
            "flats": flats,
            "all_personal_accounts": all_personal_accounts,
        }
        return JsonResponse(data, safe=False)


class DeleteHouseView(StaffRequiredMixin, DeleteView):
    def post(self, request, pk, *args, **kwargs):
        house = House.objects.get(pk=pk)
        gallery = Gallery.objects.get(pk=house.gallery_id)
        house.delete()
        gallery.delete()
        return redirect('houses')


class MailboxList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/mailbox.html'
    context_object_name = 'mailboxes'
    queryset = MailBox.objects.all()

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = SearchMessageFilterForm()
        return context


class MailboxFilteredList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/mailbox.html'
    context_object_name = 'mailboxes'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = SearchMessageFilterForm(initial=self.request.GET)
        return context

    def get_queryset(self):
        search_filter = SearchMessageFilterForm(self.request.GET)
        mailboxes = MailBox.objects.all()
        qs = []
        if search_filter.is_valid():
            if search_filter.cleaned_data['search_row']:
                qs.append(Q(description__icontains=search_filter.cleaned_data['search_row']))
                qs.append(Q(title__icontains=search_filter.cleaned_data['search_row']))
            q = Q()
            for item in qs:
                q = q | item
            mailboxes = MailBox.objects.filter(q)
        return mailboxes


class CreateMailbox(StaffRequiredMixin, CreateView):
    model = MailBox
    template_name = 'admin_panel/get_mailbox_form.html'
    form_class = MailBoxForm

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.sender_id = Personal.objects.get(user_id=self.request.user.id).id
        obj.save()
        return super().form_valid(form)

    success_url = reverse_lazy('mailboxes')


class CreateDebtorsMailbox(StaffRequiredMixin, CreateView):
    model = MailBox
    template_name = 'admin_panel/get_mailbox_form.html'

    def get(self, request, *args, **kwargs):
        form = MailBoxForm()
        form.fields['to_debtors'].initial = True
        data = {
            "form": form
        }
        return render(request, 'admin_panel/get_mailbox_form.html', context=data)

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.sender_id = Personal.objects.get(user_id=self.request.user.id).id
        obj.save()
        return super().form_valid(form)

    success_url = reverse_lazy('mailboxes')


class MailboxDetail(StaffRequiredMixin, DetailView):
    model = MailBox
    template_name = 'admin_panel/read_mailbox.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['mailbox'] = MailBox.objects.get(pk=self.kwargs['pk'])
        return context


class DeleteMailbox(StaffRequiredMixin, DeleteView):
    success_url = reverse_lazy('mailboxes')
    queryset = MailBox.objects.all()


class ApplicationList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/applications.html'
    context_object_name = 'applications'
    queryset = Application.objects.all()
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ApplicationsFilterForm()
        return context


class ApplicationFilteredList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/applications.html'
    context_object_name = 'applications'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ApplicationsFilterForm(initial=self.request.GET)
        return context

    def get_queryset(self):
        applications = Application.objects.all()
        form_filter = ApplicationsFilterForm(self.request.GET)
        qs = []
        if form_filter.is_valid():
            if form_filter.cleaned_data['number']:
                qs.append(Q(pk__icontains=form_filter.cleaned_data['number']))
            if form_filter.cleaned_data['daterange']:
                date_start, date_end = str(form_filter.cleaned_data['daterange']).split(' - ')
                date_start = date_start.split('/')
                date_end = date_end.split('/')
                date_start.reverse()
                date_end.reverse()
                date_start = "-".join(date_start)
                date_end = "-".join(date_end)
                qs.append(Q(
                    Q(date_published__gte=date_start) &
                    Q(date_published__lte=date_end)
                ))

            if form_filter.cleaned_data['master_type']:
                if form_filter.cleaned_data['master_type'] == 'any_master':
                    qs.append(Q(user_type=''))
                else:
                    qs.append(Q(user_type=form_filter.cleaned_data['master_type']))
            if form_filter.cleaned_data['description']:
                qs.append(Q(description__icontains=form_filter.cleaned_data['description']))
            if form_filter.cleaned_data['flat']:
                qs.append(Q(flat__number__icontains=form_filter.cleaned_data['flat']))
            if form_filter.cleaned_data['flat_owner']:
                full_name = str(form_filter.cleaned_data['flat_owner']).split(' ')
                qs.append(Q(
                    Q(flat__flat_owner__user__last_name__icontains=full_name[0]) &
                    Q(flat__flat_owner__patronymic__icontains=full_name[2]) &
                    Q(flat__flat_owner__user__first_name__icontains=full_name[1])
                ))
            if form_filter.cleaned_data['phone']:
                qs.append(Q(flat__flat_owner__user__phone__icontains=form_filter.cleaned_data['phone']))
            if form_filter.cleaned_data['master']:
                full_name = str(form_filter.cleaned_data['master']).split()
                qs.append(Q(
                    Q(user__user__first_name__icontains=full_name[1]) &
                    Q(user__user__last_name__icontains=full_name[0])
                ))
            if form_filter.cleaned_data['status']:
                qs.append(Q(status=form_filter.cleaned_data['status']))
            q = Q()
            for item in qs:
                q = q & item
            applications = Application.objects.filter(q)
        return applications


class ApplicationDetail(StaffRequiredMixin, DetailView):
    model = Application
    template_name = 'admin_panel/read_application.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        application = Application.objects.get(pk=self.kwargs['pk'])
        context['application'] = application
        return context


class CreateApplication(StaffRequiredMixin, CreateView):
    model = Application
    template_name = 'admin_panel/get_application_form.html'
    form_class = CreateApplicationForm
    success_url = reverse_lazy('applications')


class UpdateApplication(StaffRequiredMixin, UpdateView):
    model = Application
    template_name = 'admin_panel/update_application.html'
    form_class = ApplicationForm
    success_url = reverse_lazy('applications')


class DeleteApplication(StaffRequiredMixin, DeleteView):
    success_url = reverse_lazy('applications')
    queryset = Application.objects.all()


class CounterList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/counters.html'
    context_object_name = 'indications'
    queryset = Indication.objects.order_by('flat', 'service', '-date_published').distinct('flat', 'service')
    paginate_by = 10

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = CountersFilterForm()
        return context


class CountersFilteredList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/counters.html'
    context_object_name = 'indications'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = CountersFilterForm(initial=self.request.GET)
        return context

    def get_queryset(self):
        indications = Indication.objects.order_by('flat', 'service', '-date_published').distinct('flat', 'service')
        form_filter = CountersFilterForm(self.request.GET)
        qs = []
        if form_filter.is_valid():
            if form_filter.cleaned_data['flat']:
                qs.append(Q(flat__number=form_filter.cleaned_data['flat']))
            if form_filter.cleaned_data['house']:
                qs.append(Q(flat__house_id=form_filter.cleaned_data['house'].id))
            if form_filter.cleaned_data['section']:
                qs.append(Q(flat__section_id=form_filter.cleaned_data['section'].id))
            if form_filter.cleaned_data['service']:
                qs.append(Q(service_id=form_filter.cleaned_data['service'].id))
            q = Q()
            for item in qs:
                q = q & item
            indications = Indication.objects.order_by('flat', 'service', '-date_published').distinct('flat',
                                                                                                     'service').filter(
                q)
        return indications


class CounterIndicationsFilteredList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/counter_indications_list.html'
    context_object_name = 'indications'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['flat'] = Flat.objects.get(pk=self.kwargs['flat'])
        context['filter_form'] = CounterIndicationsFilterForm(initial=self.request.GET)
        return context

    def get_queryset(self):
        indications = Indication.objects.filter(flat_id=self.kwargs['flat'])
        form_filter = CounterIndicationsFilterForm(self.request.GET)
        qs = []
        if form_filter.is_valid():
            if form_filter.cleaned_data['house']:
                qs.append(Q(flat__house_id=form_filter.cleaned_data['house'].id))
            if form_filter.cleaned_data['section']:
                qs.append(Q(flat__section_id=form_filter.cleaned_data['section'].id))
            if form_filter.cleaned_data['service']:
                qs.append(Q(service_id=form_filter.cleaned_data['service'].id))
            if form_filter.cleaned_data['number']:
                qs.append(Q(number__icontains=form_filter.cleaned_data['number']))
            if form_filter.cleaned_data['daterange']:
                date_start, date_end = str(form_filter.cleaned_data['daterange']).split(' - ')
                date_start = date_start.split('/')
                date_end = date_end.split('/')
                date_start.reverse()
                date_end.reverse()
                date_start = "-".join(date_start)
                date_end = "-".join(date_end)
                qs.append(Q(
                    Q(date_published__gte=date_start) &
                    Q(date_published__lte=date_end)
                ))
            if form_filter.cleaned_data['status']:
                qs.append(Q(status=form_filter.cleaned_data['status']))
            q = Q()
            for item in qs:
                q = q & item
            indications = Indication.objects.filter(flat_id=self.kwargs['flat']).filter(q)
        return indications


class CounterIndicationsList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/counter_indications_list.html'
    context_object_name = 'indications'
    paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['flat'] = Flat.objects.get(pk=self.kwargs['flat'])
        context['service'] = Service.objects.get(pk=self.kwargs['service'])
        context['filter_form'] = CounterIndicationsFilterForm(initial={'service': self.kwargs['service']})

        return context

    def get_queryset(self):
        queryset = Indication.objects.filter(flat_id=self.kwargs['flat'], service_id=self.kwargs['service'])
        return queryset


class FlatIndicationsList(StaffRequiredMixin, ListView):
    template_name = 'admin_panel/counter_indications_list.html'
    context_object_name = 'indications'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        flat = Flat.objects.get(pk=self.kwargs['flat'])
        context['flat'] = Flat.objects.get(pk=self.kwargs['flat'])
        context['filter_form'] = CounterIndicationsFilterForm(initial={
            'flat': flat.number
        })

        return context

    def get_queryset(self):
        queryset = Indication.objects.filter(flat_id=self.kwargs['flat'])
        return queryset


class IndicationDetail(StaffRequiredMixin, DetailView):
    model = Indication
    template_name = 'admin_panel/read_indication.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        indication = Indication.objects.get(pk=self.kwargs['pk'])
        context['indication'] = indication
        return context


class CreateIndication(StaffRequiredMixin, CreateView):
    model = Indication
    template_name = 'admin_panel/get_indication_form.html'
    form_class = IndicationForm
    success_url = reverse_lazy('counters')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['create_new_indication'] = False
        return context

    def post(self, request, *args, **kwargs):
        indication_form = IndicationForm(request.POST)
        if indication_form.is_valid():
            obj = indication_form.save()
            if request.POST['action'] == 'save_and_new':
                indication_form = IndicationForm(initial={'indication_val': obj.indication_val, 'service': obj.service})
                data = {
                    'form': indication_form
                }
                return render(request, 'admin_panel/get_indication_form.html', context=data)
            else:
                return redirect('counters')
        else:
            data = {
                'form': indication_form
            }
            return render(request, 'admin_panel/get_indication_form.html', context=data)


class CreateNewIndication(StaffRequiredMixin, CreateView):
    model = Indication
    template_name = 'admin_panel/get_indication_form.html'
    form_class = IndicationForm

    def get_success_url(self):
        category = self.kwargs['category']  # Retrieve the category value from the URL
        return '/category/{}/'.format(category)  # Build the URL using the retrieved value

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['create_new_indication'] = True
        return context

    def get_initial(self):
        initial = super().get_initial()
        flat = Flat.objects.get(pk=self.kwargs['flat'])
        service = Service.objects.get(pk=self.kwargs['service'])
        initial['house'] = flat.house
        initial['section'] = flat.section
        initial['flat'] = flat
        initial['service'] = service
        return initial

    def post(self, request, *args, **kwargs):
        form = IndicationForm(request.POST)
        if form.is_valid():
            obj = form.save()
            if request.POST['action'] == 'save_and_new':
                indication_form = IndicationForm(initial={'indication_val': obj.indication_val, 'service': obj.service})
                data = {
                    'form': indication_form
                }
                return render(request, 'admin_panel/get_indication_form.html', context=data)
            else:
                url = reverse('counter_indications', args=[obj.flat.id, obj.service.id])
                return redirect(url)
        else:
            data = {
                'form': form,
                'create_new_indication': True,
            }
            return render(request, 'admin_panel/get_indication_form.html', context=data)


class CreateIndicationForFlat(StaffRequiredMixin, CreateView):
    model = Indication
    template_name = 'admin_panel/get_indication_form.html'
    form_class = IndicationForm

    def get_success_url(self):
        category = self.kwargs['category']  # Retrieve the category value from the URL
        return '/category/{}/'.format(category)  # Build the URL using the retrieved value

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['create_new_indication'] = True
        return context

    def get_initial(self):
        initial = super().get_initial()
        flat = Flat.objects.get(pk=self.kwargs['flat'])
        initial['house'] = flat.house
        initial['section'] = flat.section
        initial['flat'] = flat
        return initial

    def post(self, request, *args, **kwargs):
        form = IndicationForm(request.POST)
        if form.is_valid():
            obj = form.save()
            if request.POST['action'] == 'save_and_new':
                indication_form = IndicationForm(initial={'indication_val': obj.indication_val, 'service': obj.service})
                data = {
                    'form': indication_form
                }
                return render(request, 'admin_panel/get_indication_form.html', context=data)
            else:
                url = reverse('flat_indications', args=[obj.flat.id])
                return redirect(url)
        else:
            data = {
                'form': form,
                'create_new_indication': True,
            }
            return render(request, 'admin_panel/get_indication_form.html', context=data)


class UpdateIndication(StaffRequiredMixin, UpdateView):
    model = Indication
    template_name = 'admin_panel/update_indication.html'
    form_class = IndicationForm
    success_url = reverse_lazy('counters')

    def get(self, request, pk, *args, **kwargs):
        obj = Indication.objects.get(pk=pk)
        form = IndicationForm(instance=obj, initial={'house': obj.flat.house, 'section': obj.flat.section, })
        data = {
            'form': form,
        }
        return render(request, 'admin_panel/update_indication.html', context=data)


class DeleteIndication(StaffRequiredMixin, DeleteView):
    success_url = reverse_lazy('counters')
    queryset = Indication.objects.all()


from openpyxl import Workbook

# class DownloadExcel(View):
#     def get(self, request, pk, *args, **kwargs):
#         paybox = Paybox.objects.get(pk=pk)
#         wb = Workbook()
#         ws = wb.active  # это лист в excel
#         ws.title = "Выписка1"# это название листа в excel
#         wb.save('balances.xlsx')
#         data = {
#         }
#         return redirect(request.path)
